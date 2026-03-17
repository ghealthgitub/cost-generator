"""
🫚 GINGER UNIVERSE — Cost Generator
AI-powered treatment cost estimation for ginger.healthcare
Reads treatment articles, learns from real hospital pricing, generates cost sections
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from functools import wraps
import os, json, base64, io
from datetime import datetime

import config
from utils.db_connector import (
    init_cost_tables, authenticate_user, reset_user_password,
    get_specialties_list, get_treatments_by_specialty, get_treatment_by_id,
    get_hospitals_list, search_treatments,
    save_cost_source, get_recent_sources,
    save_cost_entry, save_cost_entries_batch,
    get_cost_data_for_treatment, get_calibration_data_for_specialty,
    get_all_calibration_summary,
    save_generated_cost, update_generated_cost, get_generated_cost,
    get_generation_stats, get_specialty_coverage,
    get_active_prompt, get_all_prompts, save_prompt,
    get_recent_activity
)
from utils.cost_engine import (
    generate_cost_section, quick_estimate,
    extract_from_content, extract_from_images
)
from utils.scraper import scrape_pricing_urls

# Claude API check
CLAUDE_AVAILABLE = False
try:
    from anthropic import Anthropic
    CLAUDE_AVAILABLE = bool(config.ANTHROPIC_API_KEY)
    print(f"{'✅' if CLAUDE_AVAILABLE else '⚠️'} Claude API: {'ready' if CLAUDE_AVAILABLE else 'no key set'}")
except ImportError:
    print("⚠️  anthropic package not installed")

app = Flask(__name__)
app.secret_key = config.SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# Initialize cost tables
with app.app_context():
    init_cost_tables()


# ═══════════════════════════════════════════════════════════════
# AUTH (shared user table)
# ═══════════════════════════════════════════════════════════════

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        if session['user'].get('role') != 'super_admin':
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated


@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        user = authenticate_user(email, password)
        if user:
            session['user'] = user
            return redirect(url_for('dashboard'))
        return render_template('login.html', error="Invalid credentials or insufficient permissions")
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        secret = request.form.get('secret', '').strip()
        new_pass = request.form.get('new_password', '')
        confirm_pass = request.form.get('confirm_password', '')

        if not config.RESET_SECRET:
            return render_template('reset_password.html', error="Reset not configured. Set RESET_SECRET env variable.")
        if secret != config.RESET_SECRET:
            return render_template('reset_password.html', error="Invalid reset key.")
        if not email or not new_pass:
            return render_template('reset_password.html', error="All fields are required.")
        if len(new_pass) < 6:
            return render_template('reset_password.html', error="Password must be at least 6 characters.")
        if new_pass != confirm_pass:
            return render_template('reset_password.html', error="Passwords do not match.")

        ok, msg = reset_user_password(email, new_pass)
        if ok:
            return render_template('reset_password.html', success="Password reset! You can now log in.")
        else:
            return render_template('reset_password.html', error=msg)

    return render_template('reset_password.html')


# ═══════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════

@app.route('/dashboard')
@login_required
def dashboard():
    stats = get_generation_stats()
    coverage = get_specialty_coverage()
    activity = get_recent_activity(10)
    return render_template('dashboard.html',
        user=session['user'],
        claude_available=CLAUDE_AVAILABLE,
        stats=stats,
        coverage=coverage,
        activity=activity
    )


# ═══════════════════════════════════════════════════════════════
# DATA INGESTION
# ═══════════════════════════════════════════════════════════════

@app.route('/ingest')
@login_required
def ingest_page():
    sources = get_recent_sources(20)
    return render_template('ingest.html',
        user=session['user'],
        claude_available=CLAUDE_AVAILABLE,
        sources=sources
    )


@app.route('/api/ingest/csv', methods=['POST'])
@login_required
def ingest_csv():
    """Process uploaded CSV/Excel file"""
    try:
        uploaded = request.files.get('file')
        if not uploaded or not uploaded.filename:
            return jsonify({'error': 'No file uploaded'}), 400

        fname = uploaded.filename.lower()
        content_text = ''

        if fname.endswith('.csv'):
            content_text = uploaded.read().decode('utf-8', errors='replace')
        elif fname.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), data_only=True)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows.append(f"=== Sheet: {sheet} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else '' for c in row]
                    rows.append(' | '.join(cells))
            content_text = '\n'.join(rows)
        else:
            return jsonify({'error': 'Unsupported file type. Use CSV or XLSX.'}), 400

        if not content_text.strip():
            return jsonify({'error': 'File appears to be empty'}), 400

        # Use Claude to extract structured data
        prompt_data = get_active_prompt('extraction')
        result = extract_from_content(content_text, prompt_data['prompt_text'], 'csv')

        if result.get('error'):
            return jsonify({'error': result['error']}), 500

        return jsonify({
            'success': True,
            'entries': result['entries'],
            'source_name': uploaded.filename,
            'source_type': 'csv'
        })

    except Exception as e:
        print(f"[Ingest CSV Error] {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ingest/url', methods=['POST'])
@login_required
def ingest_url():
    """Scrape hospital pricing URLs"""
    try:
        urls = request.json.get('urls', [])
        urls = [u.strip() for u in urls if u.strip()]
        if not urls:
            return jsonify({'error': 'No URLs provided'}), 400

        scraped = scrape_pricing_urls(urls)
        if not scraped['text']:
            return jsonify({'error': 'Could not extract content from any URL', 'details': scraped['errors']}), 400

        # Use Claude to extract structured data
        prompt_data = get_active_prompt('extraction')
        result = extract_from_content(scraped['text'], prompt_data['prompt_text'], 'url')

        if result.get('error'):
            return jsonify({'error': result['error']}), 500

        return jsonify({
            'success': True,
            'entries': result['entries'],
            'source_name': ', '.join(urls[:3]),
            'source_type': 'url',
            'scrape_errors': scraped['errors']
        })

    except Exception as e:
        print(f"[Ingest URL Error] {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ingest/image', methods=['POST'])
@login_required
def ingest_image():
    """Process uploaded rate card images/PDFs"""
    try:
        uploaded_files = request.files.getlist('files')
        if not uploaded_files:
            return jsonify({'error': 'No files uploaded'}), 400

        file_images = []
        for f in uploaded_files:
            if f and f.filename:
                file_data = f.read()
                if len(file_data) == 0:
                    continue

                fname = f.filename.lower()
                if fname.endswith('.png'):
                    media_type = 'image/png'
                elif fname.endswith(('.jpg', '.jpeg')):
                    media_type = 'image/jpeg'
                elif fname.endswith('.webp'):
                    media_type = 'image/webp'
                elif fname.endswith('.pdf'):
                    media_type = 'application/pdf'
                else:
                    media_type = 'image/png'

                # Resize large images
                if not fname.endswith('.pdf'):
                    try:
                        from PIL import Image
                        MAX_DIM = 7900
                        MAX_BYTES = 3.5 * 1024 * 1024
                        img = Image.open(io.BytesIO(file_data))
                        w, h = img.size
                        if w > MAX_DIM or h > MAX_DIM or len(file_data) > MAX_BYTES:
                            if w > MAX_DIM or h > MAX_DIM:
                                scale = min(MAX_DIM / w, MAX_DIM / h)
                                img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
                            if img.mode in ('RGBA', 'P', 'LA'):
                                img = img.convert('RGB')
                            media_type = 'image/jpeg'
                            for quality in [85, 70, 55, 40]:
                                buf = io.BytesIO()
                                img.save(buf, format='JPEG', quality=quality, optimize=True)
                                file_data = buf.getvalue()
                                if len(file_data) <= MAX_BYTES:
                                    break
                    except Exception as e:
                        print(f"[Resize Warning] {f.filename}: {e}")

                b64 = base64.b64encode(file_data).decode('utf-8')
                file_images.append({
                    'type': 'document' if fname.endswith('.pdf') else 'image',
                    'media_type': media_type,
                    'data': b64,
                    'filename': f.filename
                })

        if not file_images:
            return jsonify({'error': 'No valid files found'}), 400

        prompt_data = get_active_prompt('extraction')
        result = extract_from_images(file_images, prompt_data['prompt_text'])

        if result.get('error'):
            return jsonify({'error': result['error']}), 500

        return jsonify({
            'success': True,
            'entries': result['entries'],
            'source_name': ', '.join([img['filename'] for img in file_images]),
            'source_type': 'image',
            'files_processed': len(file_images)
        })

    except Exception as e:
        print(f"[Ingest Image Error] {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ingest/confirm', methods=['POST'])
@login_required
def ingest_confirm():
    """Save confirmed/reviewed extracted entries to database"""
    try:
        data = request.json
        entries = data.get('entries', [])
        source_name = data.get('source_name', 'Unknown')
        source_type = data.get('source_type', 'csv')

        if not entries:
            return jsonify({'error': 'No entries to save'}), 400

        source_id = save_cost_source(source_type, source_name, '', len(entries),
                                     session['user']['email'])
        if not source_id:
            return jsonify({'error': 'Could not create source record'}), 500

        saved = save_cost_entries_batch(entries, source_id, session['user']['email'])

        return jsonify({
            'success': True,
            'saved': saved,
            'total': len(entries),
            'source_id': source_id
        })

    except Exception as e:
        print(f"[Confirm Error] {e}")
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
# COST GENERATOR — Main screen
# ═══════════════════════════════════════════════════════════════

@app.route('/generator')
@login_required
def generator_page():
    specialties = get_specialties_list()
    return render_template('generator.html',
        user=session['user'],
        claude_available=CLAUDE_AVAILABLE,
        specialties=specialties
    )


@app.route('/api/treatments')
@login_required
def api_treatments():
    specialty_id = request.args.get('specialty_id')
    treatments = get_treatments_by_specialty(specialty_id if specialty_id else None)
    # Add cost status for each
    for t in treatments:
        gen = get_generated_cost(t['id'])
        t['cost_status'] = gen['status'] if gen else 'none'
        # Trim content for list view
        t.pop('content', None)
        t.pop('description', None)
    return jsonify(treatments)


@app.route('/api/treatments/search')
@login_required
def api_search_treatments():
    q = request.args.get('q', '')
    if len(q) < 2:
        return jsonify([])
    results = search_treatments(q)
    return jsonify(results)


@app.route('/api/generate-cost', methods=['POST'])
@login_required
def api_generate_cost():
    """Generate cost section for a treatment"""
    try:
        treatment_id = request.json.get('treatment_id')
        if not treatment_id:
            return jsonify({'error': 'No treatment specified'}), 400

        treatment = get_treatment_by_id(treatment_id)
        if not treatment:
            return jsonify({'error': 'Treatment not found'}), 404

        # Get calibration data
        cal_direct = get_cost_data_for_treatment(treatment_id=treatment_id)
        cal_specialty = get_calibration_data_for_specialty(treatment.get('specialty_name', ''))

        # Get active generation prompt
        prompt_data = get_active_prompt('generation')

        # Generate
        result = generate_cost_section(treatment, cal_direct, cal_specialty, prompt_data['prompt_text'])

        if result.get('error'):
            return jsonify({'error': result['error']}), 500

        # Save to DB
        gen_id = save_generated_cost(
            treatment_id, treatment['name'], treatment.get('specialty_name', ''),
            result['html'], result['prompt_used'], result.get('calibration_summary', ''),
            session['user']['email']
        )

        return jsonify({
            'success': True,
            'html': result['html'],
            'gen_id': gen_id,
            'treatment_name': treatment['name'],
            'had_calibration_data': len(cal_direct) > 0
        })

    except Exception as e:
        print(f"[Generate Error] {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/cost-section/<int:treatment_id>')
@login_required
def api_get_cost_section(treatment_id):
    """Get the current cost section for a treatment"""
    gen = get_generated_cost(treatment_id)
    if not gen:
        return jsonify({'exists': False})
    # Convert datetime
    for key in ('created_at', 'updated_at'):
        if gen.get(key):
            gen[key] = gen[key].isoformat()
    gen['exists'] = True
    return jsonify(gen)


@app.route('/api/cost-section/<int:gen_id>/save', methods=['POST'])
@login_required
def api_save_cost_section(gen_id):
    """Save edited cost section"""
    edited_html = request.json.get('html', '')
    status = request.json.get('status', 'draft')
    ok = update_generated_cost(gen_id, edited_html, status, session['user']['email'])
    return jsonify({'success': ok})


@app.route('/api/quick-estimate', methods=['POST'])
@login_required
def api_quick_estimate():
    """Counselor quick estimate for any procedure"""
    try:
        procedure = request.json.get('procedure', '').strip()
        if not procedure:
            return jsonify({'error': 'No procedure name provided'}), 400

        calibration = get_all_calibration_summary()
        result = quick_estimate(procedure, calibration)

        if result.get('error'):
            return jsonify({'error': result['error']}), 500

        return jsonify({
            'success': True,
            'html': result['html'],
            'procedure': procedure
        })

    except Exception as e:
        print(f"[Estimate Error] {e}")
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
# PUBLIC API — For ginger.healthcare website
# ═══════════════════════════════════════════════════════════════

@app.route('/api/v1/cost-section/<int:treatment_id>')
def public_cost_section(treatment_id):
    """Public API: returns approved HTML cost section for a treatment"""
    gen = get_generated_cost(treatment_id)
    if not gen or gen.get('status') != 'approved':
        return jsonify({'available': False}), 404
    html = gen.get('edited_html') or gen.get('generated_html', '')
    return jsonify({
        'available': True,
        'treatment_id': treatment_id,
        'treatment_name': gen.get('treatment_name', ''),
        'html': html,
        'updated_at': gen['updated_at'].isoformat() if gen.get('updated_at') else None
    })


@app.route('/api/v1/cost-estimate')
def public_cost_estimate():
    """Public API: returns AI estimate for any procedure"""
    procedure = request.args.get('q', '').strip()
    if not procedure:
        return jsonify({'error': 'Provide ?q=procedure_name'}), 400
    calibration = get_all_calibration_summary()
    result = quick_estimate(procedure, calibration)
    if result.get('error'):
        return jsonify({'error': result['error']}), 500
    return jsonify({
        'procedure': procedure,
        'html': result['html']
    })


@app.route('/api/v1/treatments/covered')
def public_covered_treatments():
    """Public API: lists treatments that have approved cost sections"""
    from utils.db_connector import get_conn
    import psycopg2.extras
    conn = get_conn()
    if not conn:
        return jsonify([])
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT cg.treatment_id, cg.treatment_name, cg.specialty_name,
                       cg.updated_at
                FROM cost_generated cg
                WHERE cg.status = 'approved'
                ORDER BY cg.treatment_name
            """)
            results = cur.fetchall()
            for r in results:
                if r.get('updated_at'):
                    r['updated_at'] = r['updated_at'].isoformat()
            return jsonify(results)
    except Exception as e:
        return jsonify([])
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# PROMPTS (admin only)
# ═══════════════════════════════════════════════════════════════

@app.route('/prompts')
@admin_required
def prompts_page():
    all_prompts = get_all_prompts()
    return render_template('prompts.html', user=session['user'], prompts=all_prompts)


@app.route('/api/prompts', methods=['GET'])
@admin_required
def api_get_prompts():
    prompts = get_all_prompts()
    for p in prompts:
        for key in ('created_at', 'updated_at'):
            if p.get(key):
                p[key] = p[key].isoformat()
    return jsonify(prompts)


@app.route('/api/prompts', methods=['POST'])
@admin_required
def api_save_prompt():
    data = request.json
    ok = save_prompt(
        data.get('id'), data.get('name', 'Untitled'),
        data.get('prompt_type', 'generation'),
        data.get('prompt_text', ''),
        data.get('set_active', False),
        session['user']['email']
    )
    return jsonify({'success': ok})


# ═══════════════════════════════════════════════════════════════
# API DOCS
# ═══════════════════════════════════════════════════════════════

@app.route('/api-docs')
@login_required
def api_docs_page():
    return render_template('api_docs.html', user=session['user'])


# ═══════════════════════════════════════════════════════════════
# HEALTH
# ═══════════════════════════════════════════════════════════════

@app.route('/health')
def health():
    return jsonify({
        'status': 'ok',
        'app': 'cost-generator',
        'version': '1.0',
        'claude_available': CLAUDE_AVAILABLE,
        'db_connected': bool(config.DATABASE_URL),
        'timestamp': datetime.now().isoformat()
    })


if __name__ == '__main__':
    app.run(debug=config.DEBUG, host='0.0.0.0', port=5001)
